import openpyxl
from pathlib import Path
path = Path('C:/Users/franc/Desktop/DIGO/2023 CONTROLE UNICO.xlsx')
wb = openpyxl.load_workbook(path, data_only=True)
print('SHEETS:', wb.sheetnames)
for name in wb.sheetnames:
    print('---', name)
    ws = wb[name]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 15:
            break
        print(i, row)

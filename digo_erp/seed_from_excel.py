import os
import openpyxl
from datetime import datetime, date
from app import create_app
from extensions import db
from models.cliente import Cliente
from models.vendedor import Vendedor
from models.materia_prima import MateriaPrima
from models.produto import ProdutoFinal
from models.venda import Venda, ContaReceber
from models.financeiro import Financeiro
from models.fornecedor import Fornecedor

def import_all():
    app = create_app()
    with app.app_context():
        print("Iniciando importação de dados da planilha em modo de baixo consumo de memória...")
        # read_only=True é muito mais rápido e consome pouquíssima memória
        wb = openpyxl.load_workbook(r'c:\Users\franc\Desktop\DIGO\2023 CONTROLE UNICO.xlsx', read_only=True, data_only=True)
        
        vendedores_set = set()
        clientes_set = set()
        
        vendas_sheets = [s for s in wb.sheetnames if s.startswith('Vendas -')]
        
        # 1. Primeiro passo: Varrer todas as vendas para colher Clientes e Vendedores únicos
        for sheet_name in vendas_sheets:
            print(f"Lendo cabeçalhos e cadastros da aba: {sheet_name}")
            ws = wb[sheet_name]
            
            empty_count = 0
            for row in ws.iter_rows(min_row=4, values_only=True):
                # row é uma tupla. Se os primeiros campos forem todos None, paramos ou pulamos
                if not row or all(c is None for c in row[:5]):
                    empty_count += 1
                    if empty_count > 15:
                        break
                    continue
                
                empty_count = 0
                
                # Vendedor está na coluna B (index 1)
                # Cliente está na coluna C (index 2)
                vendedor_nome = row[1]
                cliente_nome = row[2]
                
                if vendedor_nome and str(vendedor_nome).strip():
                    vendedores_set.add(str(vendedor_nome).strip().upper())
                if cliente_nome and str(cliente_nome).strip():
                    clientes_set.add(str(cliente_nome).strip().upper())
        
        print(f"Encontrados {len(vendedores_set)} vendedores e {len(clientes_set)} clientes na planilha.")
        
        # Cadastrar Vendedores no Banco
        vendedores_map = {}
        for v_nome in vendedores_set:
            if v_nome in ('VENDEDOR', 'PALOMA', 'UCHOA'):
                pct = 2.5
            else:
                pct = 0.0
            vendedor = Vendedor.query.filter_by(nome=v_nome).first()
            if not vendedor:
                vendedor = Vendedor(nome=v_nome, comissao_pct=pct, ativo=True)
                db.session.add(vendedor)
                db.session.flush()
            vendedores_map[v_nome] = vendedor.id
            
        # Cadastrar Clientes no Banco
        clientes_map = {}
        for c_nome in clientes_set:
            cliente = Cliente.query.filter_by(nome=c_nome).first()
            if not cliente:
                cliente = Cliente(nome=c_nome, ativo=True)
                db.session.add(cliente)
                db.session.flush()
            clientes_map[c_nome] = cliente.id
            
        db.session.commit()
        print("Clientes e Vendedores cadastrados com sucesso.")
        
        # 2. Cadastrar Fornecedores Padrão
        fornecedores_nomes = ['MC QUIMICA', 'VALE DO JUQUIA', 'VIPEL', 'FERMAFLEX', 'TRANS MEDEIROS']
        forf_map = {}
        for f_nome in fornecedores_nomes:
            f = Fornecedor.query.filter_by(nome=f_nome).first()
            if not f:
                tipo = 'materia_prima'
                if 'VIPEL' in f_nome:
                    tipo = 'embalagem'
                elif 'TRANS' in f_nome:
                    tipo = 'frete'
                f = Fornecedor(nome=f_nome, tipo=tipo, ativo=True)
                db.session.add(f)
                db.session.flush()
            forf_map[f_nome] = f.id
        db.session.commit()
        
        # Vincular fornecedor às matérias-primas
        celu = MateriaPrima.query.filter_by(tipo='celulosico').first()
        if celu and celu.fornecedor_id is None:
            celu.fornecedor_id = forf_map.get('MC QUIMICA')
        celugel = MateriaPrima.query.filter_by(tipo='celugel').first()
        if celugel and celugel.fornecedor_id is None:
            celugel.fornecedor_id = forf_map.get('VALE DO JUQUIA')
        fibra = MateriaPrima.query.filter_by(tipo='fibra').first()
        if fibra and fibra.fornecedor_id is None:
            fibra.fornecedor_id = forf_map.get('VALE DO JUQUIA')
        emb = MateriaPrima.query.filter_by(tipo='embalagem').first()
        if emb and emb.fornecedor_id is None:
            emb.fornecedor_id = forf_map.get('VIPEL')
        db.session.commit()

        # 3. Importar Vendas Históricas de forma otimizada
        produto = ProdutoFinal.query.first()
        custo_unitario = produto.custo_calculado if produto else 0.0
        vendas_count = 0
        
        for sheet_name in vendas_sheets:
            print(f"Importando registros de venda da aba: {sheet_name}")
            ws = wb[sheet_name]
            
            empty_count = 0
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or all(c is None for c in row[:5]):
                    empty_count += 1
                    if empty_count > 15:
                        break
                    continue
                
                empty_count = 0
                
                dt_val = row[0] # Data (Coluna A)
                if not dt_val:
                    continue
                
                if isinstance(dt_val, str):
                    try:
                        dt = datetime.strptime(dt_val.strip()[:10], '%Y-%m-%d').date()
                    except:
                        dt = date.today()
                elif isinstance(dt_val, datetime):
                    dt = dt_val.date()
                else:
                    dt = date.today()
                    
                v_nome = row[1] # Vendedor (Coluna B)
                c_nome = row[2] # Cliente (Coluna C)
                
                v_id = vendedores_map.get(str(v_nome).strip().upper()) if v_nome else None
                c_id = clientes_map.get(str(c_nome).strip().upper()) if c_nome else None
                
                if not c_id:
                    continue
                
                # Qtd de sacos vendidos (Coluna D - Celulose, ou Coluna H - Polímero)
                qtd_celulose = row[3]
                qtd_polimero = row[7]
                
                qtd = 0
                preco_unit = 0.0
                
                if qtd_celulose and isinstance(qtd_celulose, (int, float)):
                    qtd = int(qtd_celulose)
                    p_val = row[5] # Coluna F (Preço Celulose)
                    preco_unit = float(p_val) if p_val else 0.0
                elif qtd_polimero and isinstance(qtd_polimero, (int, float)):
                    qtd = int(qtd_polimero)
                    p_val = row[8] # Coluna I (Preço Polímero)
                    preco_unit = float(p_val) if p_val else 0.0
                
                if qtd <= 0:
                    continue
                    
                total_val = row[11] # Coluna L (Total)
                valor_liquido = float(total_val) if total_val else (qtd * preco_unit)
                
                desconto_val = row[10] # Coluna K (Desconto)
                desconto = float(desconto_val) if desconto_val else 0.0
                
                forma_pag = row[19] if len(row) > 19 else 'A VISTA'
                prazo = row[20] if len(row) > 20 else 0
                try:
                    prazo = int(prazo)
                except:
                    prazo = 0
                    
                frete_val = row[21] if len(row) > 21 else 'FOB'
                tipo_frete = 'FOB'
                if isinstance(frete_val, str) and 'CIF' in frete_val.upper():
                    tipo_frete = 'CIF'
                    
                obs = row[22] if len(row) > 22 else ''
                
                comm_pct = 2.5 if v_id else 0.0
                comissao = round(valor_liquido * comm_pct / 100, 2)
                
                custo_total = round(custo_unitario * qtd, 2)
                lucro = round(valor_liquido - custo_total, 2)
                margem = round((lucro / valor_liquido * 100) if valor_liquido else 0, 2)
                
                devendo_val = row[13] if len(row) > 13 else 0
                status_pag = 'pago'
                if devendo_val and isinstance(devendo_val, (int, float)) and devendo_val > 0:
                    status_pag = 'pendente'
                
                venda = Venda(
                    data=dt,
                    cliente_id=c_id,
                    vendedor_id=v_id,
                    quantidade_sacos=qtd,
                    preco_unitario=preco_unit,
                    desconto=desconto,
                    frete=0.0,
                    tipo_frete=tipo_frete,
                    forma_pagamento=str(forma_pag).strip().upper() if forma_pag else 'A VISTA',
                    prazo_dias=prazo,
                    valor_bruto=qtd * preco_unit,
                    valor_liquido=valor_liquido,
                    custo_total=custo_total,
                    lucro=lucro,
                    margem=margem,
                    comissao=comissao,
                    status_pagamento=status_pag,
                    valor_pago=valor_liquido if status_pag == 'pago' else 0.0,
                    observacoes=str(obs)[:200] if obs else None
                )
                db.session.add(venda)
                db.session.flush()
                
                conta = ContaReceber(
                    venda_id=venda.id,
                    cliente_id=c_id,
                    descricao=f'Venda {qtd} sacos - Planilha',
                    valor=valor_liquido,
                    vencimento=dt,
                    status=status_pag,
                    data_pagamento=dt if status_pag == 'pago' else None,
                    valor_pago=valor_liquido if status_pag == 'pago' else 0.0
                )
                db.session.add(conta)
                
                if status_pag == 'pago':
                    rec = Financeiro(
                        data=dt,
                        tipo='receita',
                        categoria='venda',
                        descricao=f'Faturamento Venda #{venda.id} - Planilha',
                        valor=valor_liquido,
                        referencia_id=venda.id,
                        referencia_tipo='venda'
                    )
                    db.session.add(rec)
                
                vendas_count += 1
                
        db.session.commit()
        print(f"Importação concluída com sucesso! Total de {vendas_count} vendas importadas.")

if __name__ == '__main__':
    import_all()
